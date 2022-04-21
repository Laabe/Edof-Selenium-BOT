from openpyxl import load_workbook
from edof.constants import EXCEL_FILE

def reset_worksheet():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.delete_rows(2, ws.max_row + 1)  # for entire sheet

    wb.save(EXCEL_FILE)
    wb.close()

def store_folder_in_excel(folder_number, status, file):
    # myFileName = '/odoo/instant_import/folder_numbers.xlsx'
    myFileName = file
    # load the workbook, and put the sheet into a variable
    wb_add = load_workbook(filename=myFileName)
    ws_add = wb_add['maj_status']

    # max_row is a sheet function that gets the last row in a sheet.
    newRowLocation = ws_add.max_row + 1
    print('new row location : ' + str(newRowLocation))

    # write to the cell you want, specifying row and column, and value :-)
    ws_add.cell(column=1, row=newRowLocation, value=str(folder_number))
    ws_add.cell(column=2, row=newRowLocation, value=status.encode('utf-8').strip())
    wb_add.save(filename=myFileName)
    wb_add.close()